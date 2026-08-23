"""SAP Excel eksportlarini o'qish.

Uch xil fayl keladi va ular BIR XIL ko'rinadi (`.xlsx`, bitta varaq,
`Sheet1`). Farqni faqat SARLAVHA qatori ko'rsatadi, shuning uchun tur
aynan shu bo'yicha aniqlanadi — fayl nomiga tayanib bo'lmaydi
(«Workbook3.xlsx», «wb1.xlsx», «savdo kunlik.xlsx» — nomlar
foydalanuvchi tomonidan har safar boshqacha yoziladi).

⚠️ BU MODULDA UCHTA TUZOQ BOR va uchalasi ham HAQIQIY ma'lumotda
o'lchangan:

  1. SUMMA IKKI AVLODDA KELADI va ular BIR XIL ustunda turadi —
     ESKI eksportda son matn (`"1 950,000"`), YANGISIDA esa oddiy
     raqam (`1230.0`). Farqni faqat katak FORMATI ko'rsatadi
     (`parse_amount` va `LegacyThousands` izohiga qarang).
  2. SANA MATN: `dd.mm.yyyy`, vaqti yo'q.
  3. TELEFON 10 XIL FORMATDA, ba'zan umuman telefon emas (`@EadTrader`).

Bu yerda BAZAGA HECH NARSA YOZILMAYDI — modul faqat o'qiydi va
tozalangan qatorlarni qaytaradi.
"""

import re
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from typing import Any

from openpyxl import load_workbook

from src.core.exceptions import AppError
from src.modules.sales.domain.entities import SaleOpType, op_type_from_sap


class SalesFileError(AppError):
    """Fayl kutilgan eksportlardan biri emas."""

    status_code = 422
    code = "sales_bad_file"


class SalesFileKind(StrEnum):
    """Qo'llab-quvvatlanadigan eksport turlari."""

    REGISTER = "register"
    """`savdo kunlik.xlsx` — operatsiyalar registri."""

    CATALOG = "catalog"
    """`Workbook3.xlsx` — kontragentlar katalogi."""

    BALANCE = "balance"
    """`Workbook1/2.xlsx` — mijoz balansi hisoboti."""


#: Ekranda ko'rsatiladigan nomlar (xato xabarlari uchun).
KIND_LABELS: dict[SalesFileKind, str] = {
    SalesFileKind.REGISTER: "savdo registri",
    SalesFileKind.CATALOG: "kontragentlar katalogi",
    SalesFileKind.BALANCE: "balans hisoboti",
}

#: Turni aniqlaydigan sarlavha ustunlari. Fayl tanilishi uchun
#: ro'yxatdagi HAMMASI bo'lishi kerak.
#
# Ataylab kam ustun tanlangan: SAP eksporti vaqt o'tib yangi ustun
# qo'shishi mumkin va bunda import yiqilmasligi kerak.
_SIGNATURES: dict[SalesFileKind, tuple[str, ...]] = {
    SalesFileKind.REGISTER: ("тип", "номер операции", "дата регистрации"),
    SalesFileKind.CATALOG: ("код бп", "название бп", "код группы"),
    SalesFileKind.BALANCE: ("kod", "klient nomi", "tel raqami"),
}

#: Sarlavha qatori shuncha birinchi qatordan qidiriladi.
#
# Amalda u har doim 1-qator, lekin SAP eksportga sarlavha (hisobot
# nomi, davr) qo'shib qo'yishi mumkin — o'shanda ham fayl tanilsin.
_HEADER_SCAN_ROWS = 5

#: Raqamning solishtiriladigan qismi. Tizimdagi boshqa joylar bilan
#: BIR XIL bo'lishi shart (`moizvonki/application/ingest.py`).
_PHONE_TAIL = 9

#: Xom matndagi `+` dan keyingi raqamlar — mamlakat kodi shu yerdan
#: olinadi. `\s*` kerak: katalogda `(+ 9989) 1234567` ko'rinishi ham bor.
_PLUS_COUNTRY = re.compile(r"\+\s*(\d+)")

#: O'zbekistonning xalqaro kodi.
_UZ_COUNTRY_CODE = "998"

#: `+CC` li raqam chet el deb hisoblanishi uchun kerakli eng kam raqam
#: soni. Bundan qisqasi — mahalliy raqam, `+` esa formatlash xatosi.
_FOREIGN_MIN_DIGITS = 11

#: SAP eksporti sonni HAR DOIM uch xona aniqlik bilan yozadi.
_DECIMALS = Decimal("0.001")

#: ESKI avlod summa katagining formati — «o'nlik ko'rsatilmagan».
#
# ⚠️ AYNAN SHU SATR ikki avlodni ajratadi, boshqa belgi yo'q.
# O'lchandi (`Хақдор ($)` ustuni bo'yicha, hamma raqam kataklar):
#
#     savdo kunlik.xlsx (eski)          649 ta → HAMMASI `#,##0`
#     клиент харакати общий (yangi)  12 591 ta → HAMMASI `General`
#     Workbook3.xlsx (eski katalog)     257 ta → `#,##0`
#     Mijozlar ruyxati.xlsx (yangi)       0 ta `#,##0` summa
#
# Yangi katalogda `#,##0` umuman uchramaydi degani EMAS: 3 ta katak
# bor, lekin uchalasi ham `Тел ракам` ustunida — u yerda summa
# o'qilmaydi, ya'ni bu belgi summa ustunlarida TOZA ajratadi.
_LEGACY_THOUSANDS_FORMAT = "#,##0"


class LegacyThousands(int):
    """ESKI eksportning noto'g'ri o'qilgan summa katagi.

    Excel `"561,000"` matnini import qilayotib vergulni MINGLIK
    ajratkichi deb hisoblagan va katakka 561 EMAS, 561000 sonini
    yozgan. Katakda `#,##0` formati aynan shundan qolgan: Excel uni
    «o'nliksiz butun son» deb belgilagan. Probelli `"1 950,000"` esa
    raqamga o'xshamagani uchun matn bo'lib qolgan — shuning uchun
    eski faylning bitta ustunida ikki xil ma'no yonma-yon yuradi.

    Tekshirildi — bunday kataklar HAR DOIM 1000 barobar katta:

        UZS hujjat:  ($) katak 8333  ↔  (сўм) matn `100 000,000`
                     8333/1000 = 8.333 $ ≈ 100 000 so'm (kurs ~12 000)
        AED hujjat:  ($) katak 136240 ↔ (дирҳам) katak 500000
                     136.240 $ ≈ 500 dirham (kurs 3.67)

    ⚠️ `int` DAN MEROS ATAYLAB. Belgi katak o'qilayotganda qo'yiladi,
    ya'ni u summa ustunidan tashqarida ham paydo bo'ladi (yangi
    katalogda — telefon raqamida). `int` bo'lib qolgani uchun
    `_text()` va boshqa hamma joy uni oddiy sondan farq qilmaydi va
    belgining ta'siri FAQAT `parse_amount` bilan cheklanadi.
    """

    __slots__ = ()


# ══════════════════════════════════════════════════════════════
#  Qiymatlarni tozalash
# ══════════════════════════════════════════════════════════════


def parse_amount(value: Any) -> Decimal | None:
    """Summani `Decimal` ga aylantiradi. `"1 950,000"` → `1950.000`.

    ⚠️ EKSPORTNING IKKI AVLODI BOR va ular bir xil ustunda keladi.
    Qaysi qoida ishlashini KATAKNING TURI hal qiladi:

      · MATN (`"1 950,000"`) — probel minglik ajratkich, vergul
        o'nlik → 1950.000. Eski eksportda summalarning ko'pi shunday.

      · `LegacyThousands` — eski eksportning Excel tomonidan buzib
        o'qilgan katagi (`#,##0` formatli butun son) → 1000 ga
        bo'linadi. Belgi `read_workbook` da qo'yiladi.

      · Oddiy RAQAM (`1230.0`, `256`, `0.0`) — YANGI eksport, qiymat
        allaqachon to'g'ri → O'ZGARISHSIZ olinadi.

    ⚠️ OXIRGI SHART QAYTA YOZILGAN. Ilgari HAR QANDAY raqam katak
    1000 ga bo'linardi va bu yangi eksportda summalarni 1000 barobar
    kichraytirib yubordi: 146 000 $ → 146 $, 256 $ → 0. Yangi faylda
    12 591 ta summa katagining hammasi raqam va hammasi `General`,
    ya'ni eski qoida ularning BARCHASINI buzardi.
    """
    if value is None or isinstance(value, bool):
        return None

    # ⚠️ `LegacyThousands` — `int`ning bolasi, shuning uchun u
    # umumiy raqam shartidan OLDIN tekshirilishi shart.
    if isinstance(value, LegacyThousands):
        return (Decimal(int(value)) / 1000).quantize(_DECIMALS)

    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(_DECIMALS)

    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text:
        return None
    # Vergul — o'nlik ajratkich. Minglik probellari yuqorida olib
    # tashlandi, ya'ni nuqta bu yerda uchramaydi.
    try:
        return Decimal(text.replace(",", ".")).quantize(_DECIMALS)
    except (InvalidOperation, ValueError):
        return None


#: Sana qanday yozilgan bo'lishi mumkin. Birinchisi — SAP dagi haqiqiy
#: format, qolganlari ehtiyot uchun (foydalanuvchi faylni Excel'da
#: ochib saqlasa format o'zgarib ketishi mumkin).
_DATE_FORMATS = ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y")


def parse_date(value: Any) -> date | None:
    """`"20.08.2026"` → `date(2026, 8, 20)`. Tanilmasa — `None`.

    ⚠️ Vaqt qismi eksportda YO'Q. Agar Excel katakni sana qilib
    o'girgan bo'lsa (`datetime`), vaqt 00:00 bo'ladi va u SOXTA —
    shuning uchun faqat sana olinadi.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def phone_key(value: Any) -> str | None:
    """Telefonning solishtiriladigan qismi — OXIRGI 9 raqam.

    SAP da 10 xil format o'lchandi: `(+99890) 1234567`, `998901234567`,
    `(90) 123-45-67`, `(+ 9989) 1234567`… Barchasi bitta kalitga
    tushadi. Ba'zi kataklarda telefon o'rniga Telegram nomi turadi
    (`@EadTrader`) — u raqamsiz qoladi va `None` qaytadi.

    9 tadan qisqa bo'lsa `None`: kalit sifatida ishonchsiz va uni
    kesib olish boshqa mijozga tushib qolish xavfini tug'diradi.

    ⚠️ «OXIRGI 9 RAQAM» QOIDASI SHARTSIZ QO'LLANMAYDI. Katalogda ikki
    turdagi qiymat noto'g'ri kalit beradi va ikkalasi ham zararli:

      · CHET EL RAQAMLARI — `(+971) …`, `(+992) …`, `(+7701) …`.
        Ularning oxirgi 9 raqami TASODIFAN o'zbek raqamiga o'xshab
        qolishi mumkin va o'shanda savdo BEGONA mijozning
        qo'ng'iroqlariga bog'lanardi. Guruh bo'yicha filtr yetarli
        EMAS: o'lchandi — 75 ta shunday qator bor, shundan 25 tasi
        «Клиенты» guruhida (qolgani asosan «Поставщики импорт»).

      · SOXTA RAQAMLAR — `(0000) 000-00-03`, `(0500) 000-00-01`,
        `(99) 999-99-99`, `(+99811) 1111111`. Ular «telefoni bor» deb
        sanaladi, keyin hech qanday qo'ng'iroq topilmaydi va mijoz
        ASOSSIZ «shubhali» bo'lib chiqadi — ya'ni aynan biz oldini
        olmoqchi bo'lgan yolg'on signal. O'lchandi: 33 ta qator
        (20 tasi noldan boshlanadi, 13 tasi bitta raqamdan iborat).

    Jami 108 kontragent kalitsiz qoladi (3531 → 3423), ya'ni qamrov
    94.3% dan 91.4% ga tushadi. Bu YO'QOTISH EMAS: o'sha 108 kalit
    baribir noto'g'ri edi — ular yo begona odamning qo'ng'iroqlariga
    olib borardi, yo hech qayerga.

    Uchala shart ATAYLAB TOR: halol raqamni o'ldirib qo'yish
    yolg'on kalitdan ham yomonroq — mijoz jimgina nazoratdan
    chiqib ketardi. Saqlanadigan chegaraviy holatlar:
    `(+99888) 8999998`, `(+99899) 5555559`, `0901234567` (trunk noli
    bilan yozilgan), shahar raqami `(71) 200-00-00`.
    """
    raw = str(value or "")
    digits = "".join(char for char in raw if char.isdigit())
    if len(digits) < _PHONE_TAIL:
        return None

    # ── 1. Chet el raqami ─────────────────────────────────────
    #
    # Xom matnda `+CC` bor va CC `998` emas. Raqam soni sharti
    # MAJBURIY: `+` bilan yozilgan qisqa mahalliy raqam
    # (`(+90) 1234567`) tasodifan chet el deb hisoblanmasin.
    # O'lchandi: `+` li 3446 qatordan 92 tasida kod 998 emas, shundan
    # 76 tasi 11 va undan ortiq raqamli — ya'ni haqiqatan xalqaro.
    match = _PLUS_COUNTRY.search(raw)
    if match is not None:
        country = match.group(1)[:3]
        if (
            len(country) == 3
            and country != _UZ_COUNTRY_CODE
            and len(digits) >= _FOREIGN_MIN_DIGITS
        ):
            return None

    tail = digits[-_PHONE_TAIL:]

    # ── 2. Noldan boshlanadi ──────────────────────────────────
    #
    # O'zbek raqamining milliy qismi HECH QACHON noldan boshlanmaydi:
    # operator kodi 33/88/90/91/93/94/95/97/98/99, shahar kodi
    # 71/62/… Noldan boshlangan «oxirgi 9 raqam» — demak qiymat
    # to'liq emas yoki umuman o'ylab topilgan.
    if tail.startswith("0"):
        return None

    # ── 3. Bitta raqamdan iborat ──────────────────────────────
    #
    # `999999999`, `111111111`, `333333333` — bo'sh katakni
    # to'ldirish uchun yozilgan qiymatlar.
    if len(set(tail)) == 1:
        return None

    return tail


def _text(value: Any, *, limit: int | None = None) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    if not text:
        return None
    return text[:limit] if limit else text


# ══════════════════════════════════════════════════════════════
#  Faylni ochish va turini aniqlash
# ══════════════════════════════════════════════════════════════


@dataclass(frozen=True, slots=True)
class SalesWorkbook:
    """O'qib olingan varaq: turi, normallashtirilgan sarlavha, qatorlar."""

    kind: SalesFileKind
    header: list[str]
    rows: list[tuple[Any, ...]]


def _normalize_header(value: Any) -> str:
    return " ".join(str(value or "").split()).lower()


def _cell_value(cell: Any) -> Any:
    """Katakning qiymati, kerak bo'lsa ESKI AVLOD belgisi bilan.

    Belgi (`LegacyThousands`) faqat ikkala shart bajarilganda qo'yiladi:
    katak formati `#,##0` VA qiymat butun son. Qolgan hamma narsa
    qiymatning o'zi bo'lib qaytadi — matn ham, `General` raqam ham,
    sana ham.

    ⚠️ `type(...) is` ATAYLAB, `isinstance` emas: `bool` ham `int`ning
    bolasi va `True` tasodifan summa deb belgilanmasligi kerak.
    Format faqat raqam katak uchun so'raladi — `number_format` har
    safar uslublar jadvaliga murojaat qiladi va uni har katak uchun
    chaqirish faylni sezilarli sekinlashtirardi.
    """
    value = cell.value
    if type(value) is int or (type(value) is float and value.is_integer()):
        if cell.number_format == _LEGACY_THOUSANDS_FORMAT:
            return LegacyThousands(value)
    return value


def _match_kind(header: Sequence[str]) -> SalesFileKind | None:
    present = set(header)
    for kind, needed in _SIGNATURES.items():
        if all(name in present for name in needed):
            return kind
    return None


def read_workbook(source: Any, *, filename: str = "") -> SalesWorkbook:
    """Faylni o'qiydi va turini SARLAVHA bo'yicha aniqlaydi.

    `source` — yo'l, bayt yoki fayl-obyekt (FastAPI `UploadFile.file`).

    ⚠️ `read_only=True` — 3746 qatorli katalog to'liq obyekt daraxtiga
    aylantirilsa yuzlab megabayt xotira ketardi. `data_only=True` —
    formulalar emas, hisoblangan qiymat kerak.

    ⚠️ QATORLAR KATAK OBYEKTI BILAN O'QILADI (`values_only` YO'Q).
    Sabab bitta: summaning qaysi avloddan ekanini faqat
    `cell.number_format` ayta oladi va u qiymatning yonida qolmaydi.
    Katak obyektlari SAQLANMAYDI — `_cell_value` darhol sof qiymat
    qaytaradi, ya'ni `read_only` ning xotira yutug'i buzilmaydi.
    O'lchandi: 12 591 qatorli faylda 0.56 s → 0.65 s.
    """
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — kutubxona har xil xato beradi
        raise SalesFileError(
            "Faylni o'qib bo'lmadi — u `.xlsx` emas yoki buzilgan. "
            "SAP dan qaytadan yuklab ko'ring."
        ) from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        kind: SalesFileKind | None = None
        header: list[str] = []
        rows: list[tuple[Any, ...]] = []
        scanned = 0

        for raw in sheet.iter_rows():
            if kind is None:
                # Sarlavhadan OLDINGI qatorlar (hisobot nomi, davr)
                # ma'lumot emas — ular hech qayerga yig'ilmaydi.
                if scanned >= _HEADER_SCAN_ROWS:
                    break
                scanned += 1
                candidate = [_normalize_header(cell.value) for cell in raw]
                kind = _match_kind(candidate)
                if kind is not None:
                    header = candidate
                continue
            values = tuple(_cell_value(cell) for cell in raw)
            if any(value is not None for value in values):
                rows.append(values)

        if kind is None:
            raise SalesFileError(
                f"«{filename or 'fayl'}» tanilmadi: sarlavhada na "
                "«Номер операции», na «Код БП», na «Kod» ustuni bor. "
                "Kutilgan fayllar: savdo registri (savdo kunlik), "
                "kontragentlar katalogi (Workbook3) yoki balans "
                "hisoboti (Workbook1/2)."
            )
        return SalesWorkbook(kind=kind, header=header, rows=rows)
    finally:
        workbook.close()


def _column(header: Sequence[str], *needles: str) -> int:
    """Sarlavhadan ustun indeksini topadi.

    ⚠️ AVVAL AYNAN TENGLIK, KEYIN QISMAN MOSLIK. Katalogda `Актив` va
    `Неактив` ustunlari yonma-yon turadi: qisman moslik `актив` ni
    `неактив` da ham topib, faol mijozni nofaol deb belgilardi.

    Qisman moslik esa boshqa joyda majburiy: `Хақдор ($)` va
    `Хақдор (cўм)` — ikkinchisida `с` LOTIN harfi (SAP eksportidagi
    haqiqiy holat), shuning uchun ustunni to'liq nom bo'yicha izlash
    mo'rt bo'lardi.
    """
    if len(needles) == 1:
        for index, name in enumerate(header):
            if name == needles[0]:
                return index
    for index, name in enumerate(header):
        if all(needle in name for needle in needles):
            return index
    raise SalesFileError(
        f"Faylda «{' + '.join(needles)}» ustuni topilmadi — eksport "
        "formati o'zgargan bo'lishi mumkin."
    )


def _cell(row: Sequence[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


# ══════════════════════════════════════════════════════════════
#  Registr — `savdo kunlik.xlsx`
# ══════════════════════════════════════════════════════════════


@dataclass(frozen=True, slots=True)
class RegisterRow:
    """Bitta operatsiya — tozalangan holda."""

    external_id: str
    doc_number: str | None
    op_type: SaleOpType
    op_type_raw: str | None
    occurred_on: date | None
    branch: str | None
    direction: str | None
    partner_code: str | None
    partner_name: str | None
    amount: Decimal | None
    amount_usd: Decimal | None
    currency: str


def parse_register(book: SalesWorkbook) -> list[RegisterRow]:
    """Registr varag'ini `RegisterRow` ro'yxatiga aylantiradi."""
    head = book.header
    col_type = _column(head, "тип")
    col_op = _column(head, "номер операции")
    col_branch = _column(head, "подразделение")
    col_direction = _column(head, "направление")
    col_doc = _column(head, "док")
    col_date = _column(head, "дата регистрации")
    col_code = _column(head, "код заказчика")
    col_name = _column(head, "название заказчика")
    # ⚠️ `Хақдор` — kompaniyaga tegishli summa (savdo, chiqim to'lovi),
    # `Қарздор` — kompaniyadan talab (kirim to'lovi, xarid). Har
    # qatorda faqat BITTA tomon to'ladi.
    col_credit_usd = _column(head, "хақдор", "$")
    col_credit_native = _column(head, "хақдор", "ў")
    col_debit_usd = _column(head, "қарздор", "$")
    col_debit_native = _column(head, "қарздор", "ў")
    col_currency = _column(head, "валюта")

    result: list[RegisterRow] = []
    for row in book.rows:
        external_id = _text(_cell(row, col_op), limit=32)
        if not external_id:
            continue

        credit_usd = parse_amount(_cell(row, col_credit_usd))
        credit_native = parse_amount(_cell(row, col_credit_native))
        debit_usd = parse_amount(_cell(row, col_debit_usd))
        debit_native = parse_amount(_cell(row, col_debit_native))

        # Qaysi tomon to'lgan bo'lsa — o'shanisi olinadi. Ikkalasi ham
        # nol bo'lsa (Бух.оп da uchraydi) `Хақдор` qoladi: summa nol
        # bo'lsa ham qator saqlanishi kerak.
        if credit_usd or credit_native:
            usd, native = credit_usd, credit_native
        elif debit_usd or debit_native:
            usd, native = debit_usd, debit_native
        else:
            usd, native = credit_usd, credit_native

        currency = _text(_cell(row, col_currency), limit=8) or "USD"
        # ⚠️ `(cўм)` ustuni aslida HUJJAT VALYUTASIDAGI summa: UZS
        # hujjatda so'm, CNY da yuan, AED da dirham. Dollar hujjatda u
        # nol bo'ladi va o'shanda `($)` ustuni hujjat summasi bo'lib
        # qoladi.
        amount = native if currency.upper() != "USD" and native else usd

        op_raw = _text(_cell(row, col_type))
        result.append(
            RegisterRow(
                external_id=external_id,
                doc_number=_text(_cell(row, col_doc), limit=32),
                op_type=op_type_from_sap(op_raw),
                op_type_raw=op_raw,
                occurred_on=parse_date(_cell(row, col_date)),
                branch=_text(_cell(row, col_branch), limit=128),
                direction=_text(_cell(row, col_direction), limit=64),
                partner_code=_text(_cell(row, col_code), limit=16),
                partner_name=_text(_cell(row, col_name), limit=255),
                amount=amount,
                amount_usd=usd,
                currency=currency.upper(),
            )
        )
    return result


# ══════════════════════════════════════════════════════════════
#  Katalog — `Workbook3.xlsx`
# ══════════════════════════════════════════════════════════════


@dataclass(frozen=True, slots=True)
class CatalogRow:
    """Bitta kontragent."""

    code: str
    name: str
    group_name: str | None
    branch: str | None
    phone: str | None
    phone_key: str | None
    is_active: bool
    telegram_link: str | None


#: `Актив` ustunidagi «ha» qiymatlari.
_YES = frozenset({"да", "yes", "ha", "true", "1", "+"})


def parse_catalog(book: SalesWorkbook) -> list[CatalogRow]:
    """Katalog varag'ini `CatalogRow` ro'yxatiga aylantiradi."""
    head = book.header
    col_name = _column(head, "название бп")
    col_code = _column(head, "код бп")
    col_group = _column(head, "код группы")
    col_phone = _column(head, "тел")
    col_branch = _column(head, "подразделение")
    col_active = _column(head, "актив")
    col_link = _column(head, "линк")

    result: list[CatalogRow] = []
    for row in book.rows:
        code = _text(_cell(row, col_code), limit=16)
        if not code:
            continue
        phone = _text(_cell(row, col_phone), limit=64)
        active = _text(_cell(row, col_active))
        result.append(
            CatalogRow(
                code=code,
                name=_text(_cell(row, col_name), limit=255) or code,
                group_name=_text(_cell(row, col_group), limit=64),
                branch=_text(_cell(row, col_branch), limit=128),
                phone=phone,
                phone_key=phone_key(phone),
                is_active=(active or "").lower() in _YES,
                telegram_link=_text(_cell(row, col_link), limit=255),
            )
        )
    return result


# ══════════════════════════════════════════════════════════════
#  Balans hisoboti — `Workbook1/2.xlsx`
# ══════════════════════════════════════════════════════════════


@dataclass(frozen=True, slots=True)
class BalanceRow:
    """Balans hisobotining bitta qatori.

    ⚠️ Bu yerda `Kod` NOYOB EMAS: qator = mijoz × filial × yo'nalish.
    Shuning uchun bu fayldan kontragent YARATILMAYDI — undan faqat
    yetishmagan telefon olinadi.
    """

    code: str
    name: str | None
    branch: str | None
    phone: str | None
    phone_key: str | None


def parse_balance(book: SalesWorkbook) -> list[BalanceRow]:
    """Balans varag'ini `BalanceRow` ro'yxatiga aylantiradi."""
    head = book.header
    col_code = _column(head, "kod")
    col_name = _column(head, "klient nomi")
    col_branch = _column(head, "bo'lim")
    col_phone = _column(head, "tel raqami")

    result: list[BalanceRow] = []
    for row in book.rows:
        code = _text(_cell(row, col_code), limit=16)
        if not code:
            continue
        phone = _text(_cell(row, col_phone), limit=64)
        result.append(
            BalanceRow(
                code=code,
                name=_text(_cell(row, col_name), limit=255),
                branch=_text(_cell(row, col_branch), limit=128),
                phone=phone,
                phone_key=phone_key(phone),
            )
        )
    return result
