"""Savdo importi — tozalash, idempotentlik, filialni biriktirish.

⚠️ HAQIQIY SAP FAYLLARIGA TAYANILMAYDI. Ular `.gitignore` da va boshqa
mashinada yo'q — testlar o'sha yerda sababsiz yiqilardi. Shuning uchun
har test o'zi KICHIK xlsx yasaydi: sarlavha haqiqiy eksportdan
so'zma-so'z ko'chirilgan (`Хақдор (cўм)` dagi `c` LOTIN harfi ham shu
jumladan), qatorlar esa faqat tekshirilayotgan holatni ko'rsatadi.

Testlar HAQIQIY dev bazasida ishlaydi, shuning uchun har biri:
  · `pytest-` prefiksli noyob kod va operatsiya raqamidan foydalanadi;
  · oxirida O'ZI yaratgan qatorlarni o'chiradi.
"""

import uuid
from collections.abc import AsyncIterator, Callable
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy import delete, func, select

from src.core.database import SessionFactory
from src.modules.agents.infrastructure.models import AgentModel
from src.modules.sales.application.importer import (
    ImportReport,
    import_catalog,
    import_file,
    import_register,
)
from src.modules.sales.application.reader import (
    LegacyThousands,
    SalesFileError,
    SalesFileKind,
    parse_amount,
    parse_date,
    parse_register,
    phone_key,
    read_workbook,
)
from src.modules.sales.domain.entities import SaleOpType, normalize_branch
from src.modules.sales.infrastructure.models import (
    SaleBranchModel,
    SaleModel,
    SalePartnerModel,
)

#: Test yozuvlarini tanib olish uchun — tozalash ham shu bo'yicha.
MARK = "pytest-"

# ── Haqiqiy eksportlardan ko'chirilgan sarlavhalar ────────────

REGISTER_HEADER = [
    "#",
    "Тип",
    "Номер операции",
    "Подразделение",
    "Направление",
    "№ док.",
    "Дата регистрации",
    "Код заказчика/поставщика",
    "Название заказчика/поставщика",
    "Название группы",
    "Хақдор ($)",
    "Хақдор (cўм)",
    "Қарздор ($)",
    "Қарздор (cўм)",
    "Валюта документа",
    "Конвертация",
]

CATALOG_HEADER = [
    "Название БП",
    "Код БП",
    "Сальдо счета",
    "Код группы",
    "Тел ракам",
    "Подразделение",
    "Лимит кредитования",
    "Лимит кредиторской задолженности",
    "Код условий оплаты",
    "Инос",
    "Актив",
    "Неактив",
    "Линк",
    "Адрес PEC",
    "ГруппаID",
    "КлентID",
]

BALANCE_HEADER = [
    "#",
    "Bo'lim",
    "Yo'nalish",
    "Kod",
    "Klient nomi",
    "Tel raqami",
    "Dastlabki qarz",
    "Oxirgi Sotuv",
    "Sotuv summa",
    "Xaftalik sotuv summa",
    "Xaftalik To'lov Summa",
    "Xaftalik tovar qaytarish summa",
    "Klientga chiqim",
    "Kurs farqi",
    "Oxirgi qarz",
]


class Legacy(int):
    """ESKI avlod katagi — xlsx ga `#,##0` formati bilan yoziladi.

    Formatning O'ZI ma'no tashiydi (`reader.LegacyThousands` izohiga
    qarang), shuning uchun uni testda qo'lda qo'yish shart: `_xlsx`
    oddiy sonni `General` qilib yozadi va o'sha katak YANGI avlod deb
    o'qiladi.
    """

    __slots__ = ()


def _xlsx(header: list[str], rows: list[list[Any]]) -> BytesIO:
    """Xotirada kichik `.xlsx` yasaydi.

    `Legacy(...)` bilan o'ralgan qiymat eski eksportning buzilgan
    katagiga aylanadi — aynan shu katakda `number_format` `#,##0`
    bo'ladi va importda 1000 ga bo'linadi.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
        for index, value in enumerate(row, start=1):
            if isinstance(value, Legacy):
                sheet.cell(row=sheet.max_row, column=index).number_format = "#,##0"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _uniq() -> str:
    return uuid.uuid4().hex[:8]


# ══════════════════════════════════════════════════════════════
#  Tozalash
# ══════════════════════════════════════════════════════════════


@pytest_asyncio.fixture
async def cleanup() -> AsyncIterator[Callable[..., None]]:
    """Test yaratgan qatorlarni oxirida o'chiradi.

    Kaskad yo'q joylar bor (`sales.agent_id` — `SET NULL`), shuning
    uchun har jadval ochiq ko'rsatiladi.
    """
    branches: set[str] = set()
    agents: set[uuid.UUID] = set()

    def _track(*, branch: str | None = None, agent: uuid.UUID | None = None) -> None:
        if branch:
            branches.add(branch)
        if agent:
            agents.add(agent)

    yield _track

    async with SessionFactory() as session:
        await session.execute(
            delete(SaleModel).where(SaleModel.external_id.like(f"{MARK}%"))
        )
        await session.execute(
            delete(SalePartnerModel).where(SalePartnerModel.code.like(f"{MARK}%"))
        )
        if branches:
            await session.execute(
                delete(SaleBranchModel).where(SaleBranchModel.branch.in_(branches))
            )
        if agents:
            await session.execute(delete(AgentModel).where(AgentModel.id.in_(agents)))
        await session.commit()


async def _make_agent(full_name: str) -> uuid.UUID:
    async with SessionFactory() as session:
        agent = AgentModel(full_name=full_name, region="Toshkent", is_active=True)
        session.add(agent)
        await session.commit()
        return agent.id


# ══════════════════════════════════════════════════════════════
#  Sof funksiyalar — bazasiz
# ══════════════════════════════════════════════════════════════


def test_matn_son_tozalanadi() -> None:
    """`"1 950,000"` — probel minglik, vergul o'nlik."""
    assert parse_amount("1 950,000") == Decimal("1950.000")
    assert parse_amount("127 867 500,000") == Decimal("127867500.000")
    assert parse_amount("0,000") == Decimal("0.000")
    assert parse_amount("") is None
    assert parse_amount(None) is None
    # Telefon yozilib qolgan katak son emas
    assert parse_amount("—") is None


def test_eski_avlod_raqam_katagi_ming_barobar_kichrayadi() -> None:
    """⚠️ Excel `"561,000"` ni 561000 deb o'qigan (`#,##0` formati).

    Vergul minglik ajratkich sifatida talqin qilingan, ya'ni katakdagi
    son 1000 barobar katta. Busiz eski faylning savdo summalari
    shishib ketardi va har qanday summa chegarasi ma'nosiz bo'lardi.
    """
    assert parse_amount(LegacyThousands(561000)) == Decimal("561.000")
    assert parse_amount(LegacyThousands(8333)) == Decimal("8.333")
    assert parse_amount(LegacyThousands(0)) == Decimal("0.000")


def test_yangi_avlod_raqam_katagi_ozgarishsiz_qoladi() -> None:
    """⚠️ YANGI eksportda qiymat TO'G'RI — unga tegilmaydi.

    Ilgari har qanday raqam katak 1000 ga bo'linardi va yangi fayl
    yuklangach summalar 1000 barobar kichrayib ketdi: 146 000 $ → 146 $,
    256 $ → 0. Yangi faylda 12 591 ta summa katagining hammasi raqam va
    hammasi `General`, ya'ni eski qoida ularning BARCHASINI buzardi.
    """
    assert parse_amount(1230.0) == Decimal("1230.000")
    assert parse_amount(2900.0) == Decimal("2900.000")

    # ⚠️ Chegaraviy holatlar — aynan foydalanuvchi ko'rgan raqamlar
    assert parse_amount(146000) == Decimal("146000.000")
    assert parse_amount(256) == Decimal("256.000")
    assert parse_amount(0.0) == Decimal("0.000")

    # Kasr qism ham saqlanadi — yangi faylda 1623 ta shunday katak bor
    assert parse_amount(1869.986) == Decimal("1869.986")


def test_sana_tahlil_qilinadi() -> None:
    assert parse_date("20.08.2026") == date(2026, 8, 20)
    assert parse_date("01.11.2024") == date(2024, 11, 1)
    assert parse_date("") is None
    assert parse_date("kecha") is None


def test_uch_xil_telefon_bitta_kalit_beradi() -> None:
    """SAP da bitta raqam 10 xil formatda yoziladi."""
    assert (
        phone_key("(+99890) 2913923")
        == phone_key("998902913923")
        == phone_key("(90) 291-39-23")
        == "902913923"
    )
    # Telegram nomi — raqam emas
    assert phone_key("@EadTrader") is None
    # 9 tadan qisqa — kalit qilib bo'lmaydi
    assert phone_key("700") is None
    assert phone_key(None) is None


def test_chet_el_raqami_kalit_bermaydi() -> None:
    """⚠️ Chet el raqamining oxirgi 9 tasi o'zbek raqamiga o'xshab qoladi.

    O'shanda savdo BEGONA mijozning qo'ng'iroqlariga bog'lanardi.
    Guruh bo'yicha filtr yetarli EMAS: o'lchandi — katalogda 75 ta
    shunday qator bor va ulardan 25 tasi «Клиенты» guruhida.
    """
    assert phone_key("(+971) 551234567") is None
    assert phone_key("(+992) 934773322") is None
    assert phone_key("(+7701) 2392507") is None
    assert phone_key("(+128) 1234567890") is None

    # ⚠️ 998 — BIZNIKI, yozilish shakli har xil bo'lsa ham tegilmaydi
    assert phone_key("(+99890) 2913923") == "902913923"
    assert phone_key("(+ 9989) 0291392") == "890291392"
    assert phone_key("(+998) 901234567") == "901234567"

    # `+` bor, lekin raqam qisqa — bu mahalliy raqam, formatlash xatosi
    assert phone_key("(+90) 1234567") == "901234567"


def test_soxta_raqam_kalit_bermaydi() -> None:
    """⚠️ Soxta raqam «telefoni bor» deb sanaladi va YOLG'ON SIGNAL beradi.

    Mijozda telefon bordek ko'rinadi, birorta qo'ng'iroq topilmaydi va
    u asossiz «shubhali» bo'lib chiqadi — aynan biz oldini olmoqchi
    bo'lgan holat. O'lchandi: katalogda 33 ta shunday qator
    (20 tasi noldan boshlanadi, 13 tasi bitta raqamdan iborat).
    """
    # Noldan boshlanadi — o'zbek raqamida bunday bo'lmaydi
    assert phone_key("(0000) 000-00-03") is None
    assert phone_key("(0500) 000-00-01") is None
    assert phone_key("(+00000) 0000018") is None
    # Hammasi bir xil raqam
    assert phone_key("(99) 999-99-99") is None
    assert phone_key("(+99811) 1111111") is None
    assert phone_key("998333333333") is None


def test_halol_chegaraviy_raqamlar_saqlanadi() -> None:
    """Filtr ATAYLAB tor — halol raqam yo'qolib ketmasin.

    Yo'qolgan raqam mijozni JIMGINA nazoratdan chiqaradi, ya'ni bu
    xato yolg'on kalitdan ham qimmatga tushadi.
    """
    # Deyarli bir xil, lekin ikki xil raqam bor
    assert phone_key("(+99888) 8999998") == "888999998"
    assert phone_key("(+99899) 5555559") == "995555559"
    # Trunk noli bilan yozilgan o'zbek raqami
    assert phone_key("0901234567") == "901234567"
    # Shahar raqami — ichida nol ko'p, lekin boshi 71
    assert phone_key("(71) 200-00-00") == "712000000"


def test_filial_nomi_imlo_farqini_yutadi() -> None:
    """`Навоий` → `Навои`, `Жиззах` → `Джиззах`."""
    assert normalize_branch("Навоий") == normalize_branch("Навои")
    assert normalize_branch("Жиззах") == normalize_branch("Джиззах")
    assert normalize_branch("  Тошкент  ") == normalize_branch("тошкент")
    # Boshqa filial bilan ARALASHIB KETMASIN
    assert normalize_branch("Тошкент") != normalize_branch("Тошкент склад")


# ══════════════════════════════════════════════════════════════
#  Fayl turini aniqlash
# ══════════════════════════════════════════════════════════════


@pytest.mark.asyncio
async def test_notogri_fayl_turi_rad_etiladi() -> None:
    """Katalog o'rniga registr berilsa — tushunarli o'zbekcha xato."""
    async with SessionFactory() as session:
        with pytest.raises(SalesFileError) as xato:
            await import_register(
                session,
                _xlsx(CATALOG_HEADER, []),
                filename="Workbook3.xlsx",
            )
    assert "kontragentlar katalogi" in str(xato.value)
    assert "savdo registri" in str(xato.value)


@pytest.mark.asyncio
async def test_begona_fayl_tanilmaydi() -> None:
    async with SessionFactory() as session:
        with pytest.raises(SalesFileError) as xato:
            await import_file(
                session, _xlsx(["Ism", "Familiya"], [["A", "B"]]), filename="xodim.xlsx"
            )
    assert "tanilmadi" in str(xato.value)


# ══════════════════════════════════════════════════════════════
#  Katalog importi
# ══════════════════════════════════════════════════════════════


@pytest.mark.asyncio
async def test_katalog_import_qilinadi(cleanup: Callable[..., None]) -> None:
    code = f"{MARK}{_uniq()}"
    book = _xlsx(
        CATALOG_HEADER,
        [
            [
                "Тест мижоз",
                code,
                "0,00000",
                "Клиенты",
                "(+99890) 2913923",
                "Бухоро",
                "0,00000",
                "0,00000",
                "4",
                None,
                "Да",
                "Нет",
                "https://t.me/test",
                None,
                None,
                None,
            ]
        ],
    )

    async with SessionFactory() as session:
        report = await import_catalog(session, book, filename="Workbook3.xlsx")

    assert report.kind is SalesFileKind.CATALOG
    assert (report.read, report.created, report.updated) == (1, 1, 0)

    async with SessionFactory() as session:
        partner = (
            await session.execute(
                select(SalePartnerModel).where(SalePartnerModel.code == code)
            )
        ).scalar_one()
    assert partner.phone_key == "902913923"
    assert partner.group_name == "Клиенты"
    assert partner.is_active is True
    assert partner.telegram_link == "https://t.me/test"


@pytest.mark.asyncio
async def test_balans_yetishmagan_telefonni_toldiradi(
    cleanup: Callable[..., None],
) -> None:
    """wb1/wb2 dan FAQAT katalogda yo'q raqam olinadi."""
    bosh = f"{MARK}{_uniq()}"
    tola = f"{MARK}{_uniq()}"

    catalog = _xlsx(
        CATALOG_HEADER,
        [
            ["Telefonsiz", bosh, "0", "Клиенты", None, "Бухоро"]
            + [None] * 4
            + ["Да", "Нет", None, None, None, None],
            ["Telefonli", tola, "0", "Клиенты", "(+99890) 1112233", "Бухоро"]
            + [None] * 4
            + ["Да", "Нет", None, None, None, None],
        ],
    )
    balance = _xlsx(
        BALANCE_HEADER,
        [
            [1, "Бухоро", "ВЕЛО", bosh, "Telefonsiz", "(+99893) 4773322"]
            + ["0,000"] * 9,
            # ⚠️ Katalogda raqami BOR mijozning boshqa raqami — u
            # o'chirib yuborilmasligi kerak
            [2, "Бухоро", "ВЕЛО", tola, "Telefonli", "(+99899) 0000000"]
            + ["0,000"] * 9,
        ],
    )

    async with SessionFactory() as session:
        await import_catalog(session, catalog, filename="wb3.xlsx")
        report = await import_catalog(session, balance, filename="wb1.xlsx")

    assert report.kind is SalesFileKind.BALANCE
    assert report.phones_filled == 1

    async with SessionFactory() as session:
        rows = dict(
            (
                await session.execute(
                    select(SalePartnerModel.code, SalePartnerModel.phone_key).where(
                        SalePartnerModel.code.in_([bosh, tola])
                    )
                )
            ).all()
        )
    assert rows[bosh] == "934773322"
    assert rows[tola] == "901112233"


# ══════════════════════════════════════════════════════════════
#  Registr importi
# ══════════════════════════════════════════════════════════════


def _register_row(
    *,
    op_number: str,
    op_type: str = "Продажа",
    branch: str = "Бухоро",
    partner_code: str,
    credit_usd: Any = "1 950,000",
    credit_uzs: Any = "0,000",
    currency: str = "USD",
    occurred: str = "10.08.2026",
) -> list[Any]:
    return [
        1,
        op_type,
        op_number,
        branch,
        "ВЕЛО",
        "89647",
        occurred,
        partner_code,
        "Тест мижоз",
        "Клиенты",
        credit_usd,
        credit_uzs,
        "0,000",
        "0,000",
        currency,
        "USD",
    ]


def test_ikki_avlod_bitta_faylda_togri_oqiladi() -> None:
    """⚠️ Summa avlodi KATAK FORMATI bo'yicha ajratiladi.

    Bu yerda haqiqiy `.xlsx` yasaladi va format ataylab qo'yiladi —
    chunki `parse_amount` ni to'g'ri chaqirish `read_workbook` ning
    ishi va aynan o'sha bog'lanish buzilgan edi. Uchala holat bitta
    faylda: eski matn, eski buzilgan katak, yangi toza raqam.
    """
    amounts = {
        # ESKI avlod — matn (probel minglik, vergul o'nlik)
        "matn": "1 950,000",
        # ESKI avlod — Excel `"561,000"` ni son deb o'qigan (`#,##0`)
        "eski-katak": Legacy(561000),
        # YANGI avlod — `General`, qiymat allaqachon to'g'ri
        "yangi": 1230.0,
        "yangi-katta": 146000,
        "yangi-kichik": 256,
        "yangi-nol": 0.0,
    }
    book = _xlsx(
        REGISTER_HEADER,
        [
            _register_row(op_number=nom, partner_code="K-1", credit_usd=summa)
            for nom, summa in amounts.items()
        ],
    )

    rows = {
        row.external_id: row.amount_usd
        for row in parse_register(read_workbook(book))
    }

    assert rows["matn"] == Decimal("1950.000")
    assert rows["eski-katak"] == Decimal("561.000")
    # ⚠️ Yangi avlodga TEGILMAYDI. Ilgari bu qatorlar 1000 ga
    # bo'linib 1.230 / 146.000 / 0.256 bo'lib qolardi.
    assert rows["yangi"] == Decimal("1230.000")
    assert rows["yangi-katta"] == Decimal("146000.000")
    assert rows["yangi-kichik"] == Decimal("256.000")
    assert rows["yangi-nol"] == Decimal("0.000")


@pytest.mark.asyncio
async def test_registr_idempotent(cleanup: Callable[..., None]) -> None:
    """Bir faylni ikki marta yuklash — nusxa qator paydo bo'lmaydi."""
    code = f"{MARK}{_uniq()}"
    op = f"{MARK}{_uniq()}"
    branch = f"{MARK}Навоий-{_uniq()}"
    cleanup(branch=branch)

    catalog = _xlsx(
        CATALOG_HEADER,
        [
            ["Тест мижоз", code, "0", "Клиенты", "(+99890) 2913923", branch]
            + [None] * 4
            + ["Да", "Нет", None, None, None, None]
        ],
    )
    async with SessionFactory() as session:
        await import_catalog(session, catalog, filename="wb3.xlsx")

    def _book() -> BytesIO:
        return _xlsx(
            REGISTER_HEADER,
            [_register_row(op_number=op, partner_code=code, branch=branch)],
        )

    async with SessionFactory() as session:
        first = await import_register(session, _book(), filename="savdo.xlsx")
    async with SessionFactory() as session:
        second = await import_register(session, _book(), filename="savdo.xlsx")

    assert (first.created, first.updated) == (1, 0)
    assert (second.created, second.updated) == (0, 1)

    async with SessionFactory() as session:
        count = (
            await session.execute(
                select(func.count()).select_from(SaleModel).where(
                    SaleModel.external_id == op
                )
            )
        ).scalar_one()
        sale = (
            await session.execute(select(SaleModel).where(SaleModel.external_id == op))
        ).scalar_one()

    assert count == 1
    assert sale.op_type == SaleOpType.SALE
    assert sale.occurred_on == date(2026, 8, 10)
    assert sale.amount == Decimal("1950.000")
    assert sale.amount_usd == Decimal("1950.000")
    # ⚠️ Telefon savdo faylida YO'Q — u katalogdan ko'chirilgan
    assert sale.phone_key == "902913923"


@pytest.mark.asyncio
async def test_filial_xodimga_biriktiriladi(cleanup: Callable[..., None]) -> None:
    """`Навоий` (SAP) → `Навои` (bizning xodim) — imlo farqi yutiladi."""
    suffix = _uniq()
    agent_name = f"{MARK}Навои-{suffix}"
    branch = f"{MARK}Навоий-{suffix}"
    yolgiz = f"{MARK}Логистика-{suffix}"
    agent_id = await _make_agent(agent_name)
    cleanup(agent=agent_id, branch=branch)
    cleanup(branch=yolgiz)

    code = f"{MARK}{_uniq()}"
    op_bor = f"{MARK}{_uniq()}"
    op_yoq = f"{MARK}{_uniq()}"

    book = _xlsx(
        REGISTER_HEADER,
        [
            _register_row(op_number=op_bor, partner_code=code, branch=branch),
            _register_row(op_number=op_yoq, partner_code=code, branch=yolgiz),
        ],
    )
    async with SessionFactory() as session:
        report = await import_register(session, book, filename="savdo.xlsx")

    # Kod katalogda yo'q — savdo baribir saqlanadi
    assert report.read == 2
    assert report.unknown_partner == 2
    assert yolgiz in report.unmatched_branches
    assert branch not in report.unmatched_branches

    async with SessionFactory() as session:
        sales = dict(
            (
                await session.execute(
                    select(SaleModel.external_id, SaleModel.agent_id).where(
                        SaleModel.external_id.in_([op_bor, op_yoq])
                    )
                )
            ).all()
        )
        rows = dict(
            (
                await session.execute(
                    select(
                        SaleBranchModel.branch, SaleBranchModel.matched_automatically
                    ).where(SaleBranchModel.branch.in_([branch, yolgiz]))
                )
            ).all()
        )

    assert sales[op_bor] == agent_id
    # ⚠️ Topilmagan filial ham `sale_branches` ga TUSHADI — rahbar uni
    # admin panelda qo'lda biriktiradi
    assert sales[op_yoq] is None
    assert rows[branch] is True
    assert rows[yolgiz] is False


@pytest.mark.asyncio
async def test_summa_hujjat_valyutasida_saqlanadi(
    cleanup: Callable[..., None],
) -> None:
    """UZS hujjatda `($)` ustuni dollar EKVIVALENTI bo'ladi."""
    code = f"{MARK}{_uniq()}"
    op = f"{MARK}{_uniq()}"
    branch = f"{MARK}Бухоро-{_uniq()}"
    cleanup(branch=branch)

    book = _xlsx(
        REGISTER_HEADER,
        [
            _register_row(
                op_number=op,
                partner_code=code,
                branch=branch,
                # Excel «8,333» ni son deb o'qigan — ESKI avlod katagi
                credit_usd=Legacy(8333),
                credit_uzs="100 000,000",
                currency="UZS",
            )
        ],
    )
    async with SessionFactory() as session:
        await import_register(session, book, filename="savdo.xlsx")
        sale = (
            await session.execute(select(SaleModel).where(SaleModel.external_id == op))
        ).scalar_one()

    assert sale.currency == "UZS"
    assert sale.amount == Decimal("100000.000")
    assert sale.amount_usd == Decimal("8.333")


@pytest.mark.asyncio
async def test_katalog_keyin_yuklansa_bogliqlik_tiklanadi(
    cleanup: Callable[..., None],
) -> None:
    """Registr avval, katalog keyin — savdo `phone_key` siz qolmasin."""
    code = f"{MARK}{_uniq()}"
    op = f"{MARK}{_uniq()}"
    branch = f"{MARK}Денов-{_uniq()}"
    cleanup(branch=branch)

    async with SessionFactory() as session:
        await import_register(
            session,
            _xlsx(
                REGISTER_HEADER,
                [_register_row(op_number=op, partner_code=code, branch=branch)],
            ),
            filename="savdo.xlsx",
        )
        sale = (
            await session.execute(select(SaleModel).where(SaleModel.external_id == op))
        ).scalar_one()
        assert sale.phone_key is None

    async with SessionFactory() as session:
        report: ImportReport = await import_catalog(
            session,
            _xlsx(
                CATALOG_HEADER,
                [
                    ["Тест мижоз", code, "0", "Клиенты", "(+99890) 2913923", branch]
                    + [None] * 4
                    + ["Да", "Нет", None, None, None, None]
                ],
            ),
            filename="wb3.xlsx",
        )

    assert report.linked_sales >= 1

    async with SessionFactory() as session:
        sale = (
            await session.execute(select(SaleModel).where(SaleModel.external_id == op))
        ).scalar_one()
    assert sale.phone_key == "902913923"


@pytest.mark.asyncio
async def test_takroriy_operatsiya_raqami_yiqitmaydi(
    cleanup: Callable[..., None],
) -> None:
    """⚠️ Bitta `Номер операции` faylda ikki marta uchraydi (o'lchangan).

    `ON CONFLICT` bitta qatorni bir so'rovda ikki marta o'zgartira
    olmaydi — busiz butun import yiqilardi.
    """
    code = f"{MARK}{_uniq()}"
    op = f"{MARK}{_uniq()}"
    branch = f"{MARK}Карши-{_uniq()}"
    cleanup(branch=branch)

    book = _xlsx(
        REGISTER_HEADER,
        [
            _register_row(op_number=op, partner_code=code, branch=branch),
            _register_row(op_number=op, partner_code=code, branch=branch),
        ],
    )
    async with SessionFactory() as session:
        report = await import_register(session, book, filename="savdo.xlsx")

    assert report.read == 2
    assert report.created == 1


@pytest.mark.asyncio
async def test_sanasiz_qator_otkazib_yuboriladi(cleanup: Callable[..., None]) -> None:
    code = f"{MARK}{_uniq()}"
    branch = f"{MARK}Нукус-{_uniq()}"
    cleanup(branch=branch)

    book = _xlsx(
        REGISTER_HEADER,
        [
            _register_row(
                op_number=f"{MARK}{_uniq()}",
                partner_code=code,
                branch=branch,
                occurred="",
            ),
            _register_row(
                op_number=f"{MARK}{_uniq()}",
                partner_code=code,
                branch=branch,
                op_type="Исходящие платежи платежи",
            ),
            _register_row(
                op_number=f"{MARK}{_uniq()}",
                partner_code=code,
                branch=branch,
                op_type="Янги тур",
            ),
        ],
    )
    async with SessionFactory() as session:
        report = await import_register(session, book, filename="savdo.xlsx")

    assert report.skipped == 1
    assert report.created == 2
    assert report.unknown_op_type == 1

    async with SessionFactory() as session:
        turlar = sorted(
            row[0]
            for row in (
                await session.execute(
                    select(SaleModel.op_type).where(
                        SaleModel.partner_code == code
                    )
                )
            ).all()
        )
    # ⚠️ SAP eksportida «Исходящие платежи» so'zi IKKI MARTA yozilgan
    assert turlar == [SaleOpType.OTHER, SaleOpType.PAYMENT_OUT]
